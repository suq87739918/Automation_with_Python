from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

wb = load_workbook("/Users/sunyueqian/Desktop/Automation with Python/pivot_table.xlsx")
sheet = wb["Report"]

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

#formulas
# sheet["B8"] = "=SUM(B6:B7)"  #we wanna work at B8, and B8 is the sum of B6 and B7
# sheet["B8"].style = "Currency"

#range
for i in range(min_column + 1,max_column + 1): #range的话end处要加1，这也是为什么这里的max_column + 1的原因，因为这个是exclusive的
    letter = get_column_letter(i) #这个会获取到这个column对应的字母，比如column = 1的话是B， column = 2 的话是C
    sheet[f'{letter}{max_row + 1}'] = f"=SUM({letter}{min_row + 1}:{letter}{max_row})"
    sheet[f"{letter}{max_row + 1}"].style = "Currency"


wb.save("report.xlsx")



