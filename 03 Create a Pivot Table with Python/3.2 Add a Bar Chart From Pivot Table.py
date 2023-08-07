from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

#file path
wb = load_workbook("/Users/sunyueqian/Desktop/Automation with Python/pivot_table.xlsx")
#sheet name
sheet = wb["Report"]

#表明需要用到的表的范围，
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

#print(min_column)  返回1，从column1开始
#print(max_column)  返回7，到column7结束
#print(min_row)     返回5
#print(max_row)     返回7

barchart = BarChart()

#min_col需要加1因为真实数据是从第二个column开始的，第一个column里的gender信息
data = Reference(sheet,
                 min_col=min_column+1,
                 max_col=max_column,
                 min_row=min_row,
                 max_row=max_row)  # including headers
#min_row需要加1因为真实有效的分类信息是从第6个row开始的，第5个row不是分类
categories = Reference(sheet,
                       min_col=min_column,
                       max_col=min_column,
                       min_row=min_row+1,
                       max_row=max_row)  # not including headers

#create barchart
# Adding data and categories
barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

# Make chart
sheet.add_chart(barchart, "B12")
barchart.title = 'Sales by Product line'
barchart.style = 5  # choose the chart style
wb.save("barchart.xlsx")

